"""
Smart OSINT Lookup Web App (Fixed Full Version)
"""

import os, json, datetime
from flask import Flask, render_template, request, redirect, session, send_file, flash, abort
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from werkzeug.security import generate_password_hash, check_password_hash

# ---------- CONFIG ----------
APP_SECRET = os.environ.get("OSINT_SECRET", "change_this_secret")
ADMIN_USER = os.environ.get("OSINT_ADMIN_USER", "admin")
ADMIN_PASS = os.environ.get("OSINT_ADMIN_PASS", "admin123")
FREE_LIMIT = int(os.environ.get("OSINT_FREE_LIMIT", "2"))
IP_ADDRESS = os.environ.get("OSINT_IP", "0.0.0.0")
PORT = int(os.environ.get("OSINT_PORT", "8000"))

DATA_FILE = "data.json"
DATA_DIR = "data"
USER_FILE = "users.json"
LOG_FILE = "logs.json"
TEMPLATE_DIR = "templates"
STATIC_DIR = "static"

# ---------- SAFE PRINT (no emoji) ----------
def log(msg):
    try:
        print(msg)
    except Exception:
        pass

# ---------- INITIAL SETUP ----------
def ensure_dirs_and_templates():
    os.makedirs(TEMPLATE_DIR, exist_ok=True)
    os.makedirs(STATIC_DIR, exist_ok=True)
    if not os.path.exists(os.path.join(STATIC_DIR, "style.css")):
        with open(os.path.join(STATIC_DIR, "style.css"), "w", encoding="utf-8") as f:
            f.write("body{background:#f6f8fb} .card{border-radius:12px}")

def ensure_data_files():
    os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(DATA_FILE):
        sample = [
            {"name": "John Doe", "email": "john@example.com", "phone": "9876543210", "city": "Delhi"},
            {"name": "Alice Kumar", "email": "alice@domain.com", "phone": "9998887777", "city": "Mumbai"}
        ]
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(sample, f, indent=2)
    if not os.path.exists(USER_FILE):
        with open(USER_FILE, "w", encoding="utf-8") as f:
            json.dump({}, f)
    if not os.path.exists(LOG_FILE):
        with open(LOG_FILE, "w", encoding="utf-8") as f:
            json.dump([], f)

# ---------- JSON HELPERS ----------
def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)

# ---------- SEARCH ----------
import csv

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

def _normalize_json_payload(raw):
    if isinstance(raw, list):
        return raw
    if isinstance(raw, dict):
        if isinstance(raw.get("data"), list):
            return raw["data"]
        out = []
        for v in raw.values():
            if isinstance(v, list):
                out.extend(v)
        return out
    return []

def _iter_records_from_json(path):
    try:
        raw = load_json(path)
    except Exception as e:
        log(f"JSON load failed for {path}: {e}")
        return
    for item in _normalize_json_payload(raw):
        if isinstance(item, str):
            try:
                item = json.loads(item)
            except json.JSONDecodeError:
                continue
        if isinstance(item, dict):
            yield item

def _iter_records_from_csv(path):
    try:
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if isinstance(row, dict):
                    yield row
    except Exception as e:
        log(f"CSV read failed for {path}: {e}")

def _iter_records_from_xlsx(path):
    if not load_workbook:
        log(f"openpyxl not installed; skipping .xlsx file: {path}")
        return
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active or wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return
        header = [str(h).strip() if h is not None else f"col{i+1}" for i, h in enumerate(header)]
        for r in rows:
            item = {header[i]: (r[i] if i < len(r) else None) for i in range(len(header))}
            yield item
    except Exception as e:
        log(f"XLSX read failed for {path}: {e}")

COMMON_FIELDS = ("name", "email", "phone", "mobile", "address", "city", "fname", "circle", "id", "alt")

def _record_matches(item, q):
    for k in COMMON_FIELDS:
        if q in str(item.get(k, "")).lower():
            return True
    for v in item.values():
        if q in str(v).lower():
            return True
    return False

def _iter_all_records():
    if os.path.exists(DATA_FILE):
        for item in _iter_records_from_json(DATA_FILE):
            yield item
    if os.path.isdir(DATA_DIR):
        try:
            for name in sorted(os.listdir(DATA_DIR)):
                path = os.path.join(DATA_DIR, name)
                if not os.path.isfile(path):
                    continue
                lower = name.lower()
                if lower.endswith('.json'):
                    yield from _iter_records_from_json(path)
                elif lower.endswith('.csv'):
                    yield from _iter_records_from_csv(path)
                elif lower.endswith('.xlsx') or lower.endswith('.xlsm'):
                    yield from _iter_records_from_xlsx(path)
        except Exception as e:
            log(f"Error reading data dir: {e}")

def search_data(query):
    q = (query or "").strip().lower()
    if not q:
        return []
    results = []
    seen = set()
    for item in _iter_all_records():
        try:
            if _record_matches(item, q):
                sig = json.dumps(item, sort_keys=True, default=str)
                if sig not in seen:
                    seen.add(sig)
                    results.append(item)
        except Exception as e:
            log(f"Record processing error: {e}")
            continue
    return results

# ---------- PDF ----------
def create_pdf_bytes(username, query, results):
    buf = BytesIO()
    p = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    margin = 40
    y = height - margin

    p.setFont("Helvetica-Bold", 16)
    p.drawString(margin, y, "Smart OSINT Lookup - Search Result")
    y -= 30
    p.setFont("Helvetica", 10)
    p.drawString(margin, y, f"User: {username} | Query: {query} | Time: {datetime.datetime.utcnow().isoformat()}Z")
    y -= 20
    p.line(margin, y, width - margin, y)
    y -= 20

    if not results:
        p.setFont("Helvetica-Oblique", 12)
        p.drawString(margin, y, "No results found.")
    else:
        for idx, r in enumerate(results, start=1):
            if y < 120:
                p.showPage()
                y = height - margin
            p.setFont("Helvetica-Bold", 12)
            p.drawString(margin, y, f"Result {idx}")
            y -= 16
            p.setFont("Helvetica", 10)
            for k, v in r.items():
                text = f"{k.capitalize()}: {v}"
                p.drawString(margin + 10, y, text)
                y -= 14
            p.line(margin, y, width - margin, y)
            y -= 12

    p.showPage()
    p.save()
    buf.seek(0)
    return buf

# ---------- FLASK APP ----------
app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR)
app.secret_key = APP_SECRET

ensure_dirs_and_templates()
ensure_data_files()

# ---------- ROUTES ----------
@app.route("/", methods=["GET"])
def index():
    return render_template("login.html", message=None, free_limit=FREE_LIMIT)

@app.route("/login", methods=["POST"])
def login():
    username = request.form.get("username", "").strip().lower()
    password = request.form.get("password", "")
    if not username or not username.isalnum():
        return render_template("login.html", message="Enter a valid alphanumeric username.", free_limit=FREE_LIMIT)
    if not password or len(password) < 6:
        return render_template("login.html", message="Password must be at least 6 characters.", free_limit=FREE_LIMIT)

    users = load_json(USER_FILE)
    user = users.get(username)

    if user is None:
        users[username] = {"count": 0, "limit": FREE_LIMIT, "last_search": None, "password": generate_password_hash(password)}
        save_json(USER_FILE, users)
    else:
        if "password" not in user:
            user["password"] = generate_password_hash(password)
            users[username] = user
            save_json(USER_FILE, users)
        elif not check_password_hash(user.get("password", ""), password):
            return render_template("login.html", message="Invalid username or password.", free_limit=FREE_LIMIT)

    session["user"] = username
    return redirect("/home")

@app.route("/home", methods=["GET", "POST"])
def home():
    if "user" not in session:
        return redirect("/")
    username = session["user"]
    users = load_json(USER_FILE)
    user = users.get(username, {"count": 0, "limit": FREE_LIMIT, "last_search": None})
    message = None
    results = None
    query = None

    if request.method == "POST":
        query = request.form.get("query", "").strip()
        log(f"Searching for: {query}")
        if not query:
            message = "Enter a search term."
        else:
            if user["count"] >= user["limit"]:
                log("Search limit reached")
                return redirect("/upgrade")

            results = search_data(query)
            log(f"Found {len(results)} result(s)")
            user["count"] += 1
            user["last_search"] = datetime.datetime.utcnow().isoformat()
            users[username] = user
            save_json(USER_FILE, users)

            logs = load_json(LOG_FILE)
            logs.append({"time": datetime.datetime.utcnow().isoformat(), "user": username, "query": query, "count": len(results)})
            save_json(LOG_FILE, logs)

    return render_template("home.html", username=username, user=user, results=results, message=message, query=query)

@app.route("/download_pdf")
def download_pdf():
    if "user" not in session:
        return redirect("/")
    q = request.args.get("q", "").strip()
    if not q:
        flash("No query provided")
        return redirect("/home")
    username = session["user"]
    results = search_data(q)
    pdf_bytes = create_pdf_bytes(username, q, results)
    fname = f"result_{username}_{datetime.datetime.utcnow().strftime('%Y%m%d%H%M%S')}.pdf"
    return send_file(pdf_bytes, as_attachment=True, download_name=fname, mimetype="application/pdf")

@app.route("/upgrade", methods=["GET"])
def upgrade():
    if "user" not in session:
        return redirect("/")
    username = session["user"]
    users = load_json(USER_FILE)
    user = users.get(username, {"count": 0, "limit": FREE_LIMIT})
    return render_template("upgrade.html", user=user)

@app.route("/request_upgrade", methods=["POST"])
def request_upgrade():
    if "user" not in session:
        return redirect("/")
    plan = int(request.form.get("plan", "0"))
    logs = load_json(LOG_FILE)
    logs.append({"time": datetime.datetime.utcnow().isoformat(), "user": session["user"], "query": f"REQUEST_UPGRADE_{plan}", "count": 0})
    save_json(LOG_FILE, logs)
    return render_template("upgrade.html", user=load_json(USER_FILE).get(session["user"]))

@app.route("/admin", methods=["GET", "POST"])
def admin():
    if request.method == "GET":
        if session.get("admin"):
            users = load_json(USER_FILE)
            logs = load_json(LOG_FILE)
            logs = sorted(logs, key=lambda x: x.get("time", ""), reverse=True)[:200]
            return render_template("admin.html", login=False, users=users, logs=logs)
        return render_template("admin.html", login=True, message=None)

    if not session.get("admin"):
        return abort(403)

    username = request.form.get("username")
    try:
        add_limit = int(request.form.get("add_limit", "0"))
    except Exception:
        add_limit = 0

    users = load_json(USER_FILE)
    if username in users and add_limit > 0:
        users[username]["limit"] = users[username].get("limit", FREE_LIMIT) + add_limit
        save_json(USER_FILE, users)
    return redirect("/admin")

@app.route("/admin/login", methods=["POST"])
def admin_login():
    user = request.form.get("username")
    pwd = request.form.get("password")
    if user == ADMIN_USER and pwd == ADMIN_PASS:
        session["admin"] = True
        return redirect("/admin")
    return render_template("admin.html", login=True, message="Invalid credentials")

@app.route("/admin/logout")
def admin_logout():
    session.pop("admin", None)
    return redirect("/admin")

@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect("/")

# ---------- RUN ----------
if __name__ == "__main__":
    log(f"Running Smart OSINT Web App at http://{IP_ADDRESS}:{PORT}")
    app.run(host=IP_ADDRESS, port=PORT, debug=False)
